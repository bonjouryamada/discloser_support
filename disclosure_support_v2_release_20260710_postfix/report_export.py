from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_service import regime_text
from disclosure_calculator import extract_summary_lines, inject_dynamic_borderlines


NAVY = "1F4E79"
BLUE = "4472C4"
PALE_YELLOW = "FFF2CC"
PALE_ORANGE = "F8CBAD"


def _summary(text: str, financial_data: dict[str, float]) -> str:
    rows = extract_summary_lines(text, financial_data)
    if not rows:
        return "該当する定量的な数値判定ラインはありません。"
    return "\n".join(f"・{row['borderline']}" for row in rows)


def generate_report(
    record: dict[str, Any],
    financial_data: dict[str, float],
    company_query: str = "",
    fetched_company_name: str = "",
    edinet_doc_info: dict[str, Any] | None = None,
) -> bytes:
    doc_info = edinet_doc_info or financial_data.get("doc_info") or {}
    document_name = doc_info.get("doc_description") or doc_info.get("document_title") or ""
    submit_date = str(doc_info.get("submit_datetime") or "")[:10]
    period_label = doc_info.get("period_label") or ""
    doc_id = doc_info.get("doc_id") or ""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "判定支援レポート"
    sheet.freeze_panes = "A10"

    sheet.merge_cells("A1:D2")
    sheet["A1"] = "制度横断 開示判定支援レポート"
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

    metadata = [
        ("会社名・証券コード入力", company_query),
        ("EDINET取得企業名", fetched_company_name),
        ("EDINET書類名", document_name),
        ("EDINET提出日", submit_date),
        ("EDINET対象決算期", period_label),
        ("EDINET docID", doc_id),
        ("開示区分", record["disclosure_category"]),
        ("開示項目", record["disclosure_item"]),
        ("手動レビュー", "要" if record["manual_review_flag"] else "不要"),
        ("手動レビュー理由", record["manual_review_reason"]),
        ("PDF頁", ", ".join(map(str, record.get("pdf_pages", [])))),
    ]
    for row_number, (label, value) in enumerate(metadata, start=4):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, value)
        sheet.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=4)

    start = 4 + len(metadata) + 2
    headers = ["制度", "数値ライン要約", "条文・基準本文", "注意"]
    for column, value in enumerate(headers, start=1):
        cell = sheet.cell(start, column, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    regimes = [
        ("東証適時開示", regime_text(record, "timely")),
        ("内部者取引規制", regime_text(record, "insider")),
        ("臨時報告書", regime_text(record, "extraordinary")),
    ]
    for row_number, (label, text) in enumerate(regimes, start=start + 1):
        sheet.cell(row_number, 1, label)
        sheet.cell(row_number, 2, _summary(text, financial_data))
        sheet.cell(row_number, 3, inject_dynamic_borderlines(text, financial_data))
        sheet.cell(row_number, 4, "法的判断・提出要否は公式一次情報と個別事実で確認してください。")
        sheet.row_dimensions[row_number].height = 150

    if record["manual_review_flag"]:
        for cell in sheet[start + 1 : start + 4]:
            for item in cell:
                item.fill = PatternFill("solid", fgColor=PALE_ORANGE)

    financial_row = start + 6
    sheet.cell(financial_row, 1, "財務数値（百万円）")
    sheet.cell(financial_row, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.cell(financial_row, 1).font = Font(color="FFFFFF", bold=True)
    if period_label:
        sheet.cell(financial_row, 2, f"対象決算期: {period_label}")
        sheet.merge_cells(start_row=financial_row, start_column=2, end_row=financial_row, end_column=4)

    metrics = [
        ("連結純資産", "net_assets"),
        ("連結売上高", "net_sales"),
        ("連結経常利益", "recurring_profit"),
        ("連結純利益", "net_income"),
        ("資本金", "capital_stock"),
    ]
    for offset, (label, key) in enumerate(metrics, start=1):
        sheet.cell(financial_row + offset, 1, label)
        sheet.cell(financial_row + offset, 2, financial_data.get(key, 0))
        sheet.cell(financial_row + offset, 2).number_format = "#,##0"

    disclaimer_row = financial_row + len(metrics) + 2
    sheet.cell(disclaimer_row, 1, "免責")
    sheet.cell(disclaimer_row, 1).fill = PatternFill("solid", fgColor=BLUE)
    sheet.cell(disclaimer_row, 1).font = Font(color="FFFFFF", bold=True)
    sheet.merge_cells(
        start_row=disclaimer_row,
        start_column=2,
        end_row=disclaimer_row + 1,
        end_column=4,
    )
    sheet.cell(disclaimer_row, 2, (
        "本資料は判定補助情報です。法的判断を確定せず、最新の法令・JPX規則・"
        "個別事実を必ず確認してください。"
    ))
    sheet.cell(disclaimer_row, 2).fill = PatternFill("solid", fgColor=PALE_YELLOW)

    widths = [22, 45, 90, 36]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
