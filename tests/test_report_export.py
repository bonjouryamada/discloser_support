from io import BytesIO
import unittest

from openpyxl import load_workbook

from data_service import load_records, regime_text
from report_export import generate_report


class ReportExportTests(unittest.TestCase):
    def _cell_value_by_label(self, sheet, label):
        for row in sheet.iter_rows():
            if row[0].value == label:
                return row[1].value
        self.fail(f"label not found in report: {label}")

    def _labels(self, sheet):
        return [row[0].value for row in sheet.iter_rows() if row[0].value]

    def _row_number_by_label(self, sheet, label):
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            if row[0].value == label:
                return row_number
        self.fail(f"label not found in report: {label}")

    def test_report_can_be_opened(self):
        financial_data = {
            "net_assets": 1_000,
            "net_sales": 2_000,
            "recurring_profit": 100,
            "net_income": 80,
            "capital_stock": 500,
        }
        payload = generate_report(load_records()[0], financial_data, "1234", "テスト株式会社")
        workbook = load_workbook(BytesIO(payload), read_only=True)
        self.assertEqual(workbook.sheetnames, ["判定支援レポート"])
        self.assertEqual(workbook["判定支援レポート"]["A1"].value, "制度横断 開示判定支援レポート")

    def test_report_contains_required_business_fields(self):
        financial_data = {
            "net_assets": 1_000,
            "net_sales": 2_000,
            "recurring_profit": 100,
            "net_income": 80,
            "capital_stock": 500,
            "doc_info": {
                "doc_description": "有価証券報告書",
                "submit_datetime": "2026-06-27 15:00",
                "period_label": "2026年3月期",
                "doc_id": "S100TEST",
            },
        }
        record = next(item for item in load_records() if item["mapping_id"] == "LC-DEC-13")

        payload = generate_report(
            record,
            financial_data,
            company_query="7203",
            fetched_company_name="サンプル株式会社",
        )
        sheet = load_workbook(BytesIO(payload), read_only=True)["判定支援レポート"]

        self.assertEqual(self._cell_value_by_label(sheet, "会社名・証券コード入力"), "7203")
        self.assertEqual(self._cell_value_by_label(sheet, "EDINET取得企業名"), "サンプル株式会社")
        self.assertEqual(self._cell_value_by_label(sheet, "EDINET書類名"), "有価証券報告書")
        self.assertEqual(self._cell_value_by_label(sheet, "EDINET提出日"), "2026-06-27")
        self.assertEqual(self._cell_value_by_label(sheet, "EDINET対象決算期"), "2026年3月期")
        self.assertEqual(self._cell_value_by_label(sheet, "EDINET docID"), "S100TEST")
        self.assertEqual(self._cell_value_by_label(sheet, "開示区分"), record["disclosure_category"])
        self.assertEqual(self._cell_value_by_label(sheet, "開示項目"), record["disclosure_item"])
        self.assertNotIn("mapping_id", self._labels(sheet))
        self.assertEqual(self._cell_value_by_label(sheet, "手動レビュー"), "要")
        self.assertIn("一対多", self._cell_value_by_label(sheet, "手動レビュー理由"))
        self.assertEqual(self._cell_value_by_label(sheet, "PDF頁"), "8, 9, 10")

        criteria_header_row = self._row_number_by_label(sheet, "制度")
        expected_regimes = [
            ("東証適時開示", "timely"),
            ("内部者取引規制", "insider"),
            ("臨時報告書", "extraordinary"),
        ]
        for row_number, (label, regime_key) in enumerate(expected_regimes, start=criteria_header_row + 1):
            self.assertEqual(sheet.cell(row_number, 1).value, label)
            self.assertIn("判定ライン:", sheet.cell(row_number, 2).value)
            self.assertIn("判定ライン:", sheet.cell(row_number, 3).value)
            self.assertIn(regime_text(record, regime_key).split()[0], sheet.cell(row_number, 3).value)

        financial_row = self._row_number_by_label(sheet, "財務数値（百万円）")
        self.assertEqual(sheet.cell(financial_row, 2).value, "対象決算期: 2026年3月期")
        expected_metrics = [
            ("連結純資産", 1_000),
            ("連結売上高", 2_000),
            ("連結経常利益", 100),
            ("連結純利益", 80),
            ("資本金", 500),
        ]
        for row_offset, (label, value) in enumerate(expected_metrics, start=1):
            self.assertEqual(sheet.cell(financial_row + row_offset, 1).value, label)
            self.assertEqual(sheet.cell(financial_row + row_offset, 2).value, value)

        disclaimer_row = self._row_number_by_label(sheet, "免責")
        self.assertIn("法的判断を確定せず", sheet.cell(disclaimer_row, 2).value)


if __name__ == "__main__":
    unittest.main()
