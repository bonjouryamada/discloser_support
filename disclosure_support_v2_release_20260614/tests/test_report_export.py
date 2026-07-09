from io import BytesIO
import unittest

from openpyxl import load_workbook

from data_service import load_records
from report_export import generate_report


class ReportExportTests(unittest.TestCase):
    def test_report_can_be_opened(self):
        financial_data = {
            "net_assets": 1_000,
            "net_sales": 2_000,
            "recurring_profit": 100,
            "net_income": 80,
            "capital_stock": 500,
        }
        payload = generate_report(load_records()[0], financial_data, "1234", "テスト株式会社")
        workbook = load_workbook(BytesIO(payload), read_only=True)
        self.assertEqual(workbook.sheetnames, ["判定支援レポート"])
        self.assertEqual(workbook["判定支援レポート"]["A1"].value, "制度横断 開示判定支援レポート")


if __name__ == "__main__":
    unittest.main()
